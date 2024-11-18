import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Step 1: Create the DataFrame and save it to an Excel file
data = {
    'Name': ['Alice', 'Bob', 'Charlie', 'David'],
    'Age': [24, 30, 35, 40],
    'Positive': [True, False, True, False],  # True/False column where we want checkboxes
    'Negative': [False, True, False, True]   # Another True/False column for checkboxes
}
df = pd.DataFrame(data)
output_file = 'spreadsheet_with_checkboxes.xlsx'
df.to_excel(output_file, index=False)

# Step 2: Open the Excel file with openpyxl
wb = load_workbook(output_file)
ws = wb.active

# Step 3: Define a function to add checkboxes based on True/False values in DataFrame
def add_checkboxes(df, column_name, ws):
    # Find the column letter in Excel for the specified column name
    col_letter = chr(65 + df.columns.get_loc(column_name))
    start_row = 2  # assuming header is in the first row

    # Add checkboxes (data validation dropdowns) for each row in the column
    for row in range(start_row, start_row + len(df)):
        # Create a Data Validation object for a dropdown with "TRUE" and "FALSE" options
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        ws.add_data_validation(dv)
        
        # Set initial value based on DataFrame value
        cell = ws[f"{col_letter}{row}"]
        cell.value = "TRUE" if df[column_name].iloc[row - start_row] else "FALSE"
        dv.add(cell)

# Step 4: Add checkboxes to 'Positive' and 'Negative' columns
add_checkboxes(df, 'Positive', ws)
add_checkboxes(df, 'Negative', ws)

# Step 5: Save the updated Excel file
wb.save(output_file)

print(f"Spreadsheet with checkboxes saved as {output_file}.")
